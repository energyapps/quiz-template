from openpyxl import Workbook
from openpyxl import load_workbook
from string import ascii_uppercase

data = load_workbook('data.xlsx')
headlines = load_workbook('headlines.xlsx')

columnheads = headlines['Sheet1']
questions = data['Questions']
headerinfo = data['Header Info']
footerinfo = data['Footer Info']

# Create three workbooks, i.e. 3 XLS files.
wb1 = Workbook()
wb2 = Workbook()
wb3 = Workbook()

#### ADD sheet 1 (Questions) to the first work book

# grab the active worksheet
ws1 = wb1.active

# Add the headline
for c in ascii_uppercase:
	i = c + "1"
	ws1[i] = columnheads[i].value

#Match the first tab with the wb1, starting at 2 to not add in the header line, 
#which we'll switch in from a different file, for safer JSON outputs
for x in range(2, 100):
	o = "A" + str(x)
	if questions[o].value != None:
		for c in ascii_uppercase:
			i = c + str(x)

			ws1[i] = questions[i].value

wb1.save("wb1.xlsx")

#### ADD sheet 2 (Header info) to the second work book
ws2 = wb2.active

for c in ascii_uppercase:
	ws2[c + "1"] = columnheads[c + "2"].value

for x in range(2, 100):
	for c in ascii_uppercase:
		i = c + str(x)
		ws2[i] = headerinfo[i].value

wb2.save("wb2.xlsx")

#### ADD sheet 3 (Header info) to the third work book

ws3 = wb3.active

for c in ascii_uppercase:
	ws3[c + "1"] = columnheads[c + "3"].value

for x in range(3, 100):
	for c in ascii_uppercase:
		i = c + str(x)
		j = c + str(x - 1)
		ws3[j] = footerinfo[i].value

wb3.save("wb3.xlsx")
